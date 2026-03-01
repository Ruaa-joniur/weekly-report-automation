import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties


BASE_DIR = Path(__file__).resolve().parent

CLOSED_FILE = BASE_DIR / "closed call report.xlsx"
OPEN_FILE = BASE_DIR / "open call report.xlsx"
USER_MISSING_FILE = BASE_DIR / "user missing agency.xlsx"
OUTPUT_FILE = BASE_DIR / "EUS & SD Weekly Report.xlsx"

# اسم العمود المطلوب فلترته (كما في الملف الخام)
ASSIGNED_GROUP_COL = "Assigned Group*+"

# أسماء الأعمدة في الملف
ORG_COL = "Organization"
DEPT_COL = "Department"
FIRST_NAME_COL = "First Name+"
LAST_NAME_COL = "Last Name+"

# صف الهيدر في ملف الإخراج
HEADER_ROW = 1
DATA_START_ROW = HEADER_ROW + 1

# القيم المسموح بها في عمود Assigned Group*+
ALLOWED_GROUPS = {
    "SCCM-Nexthink Support",
    "Service Desk",
    "OnSite  EUS",  # انتبه للمسافة المزدوجة كما في النص
}


def _normalize_col_name(name: str) -> str:
    # نطبع الاسم بحيث يتجاهل المسافات والرموز مثل * +
    return "".join(ch.lower() for ch in str(name) if ch.isalnum())


def _find_assigned_group_col(df: pd.DataFrame) -> str:
    # أولاً: تطابق مباشر
    if ASSIGNED_GROUP_COL in df.columns:
        return ASSIGNED_GROUP_COL

    # ثانياً: تطابق بعد التطبيع لتجاوز اختلاف الرموز/المسافات
    target = _normalize_col_name(ASSIGNED_GROUP_COL)
    for col in df.columns:
        if _normalize_col_name(col) == target:
            return col

    raise KeyError(f"Column '{ASSIGNED_GROUP_COL}' not found in the input file.")


def filter_assigned_group(df: pd.DataFrame) -> pd.DataFrame:
    """الإبقاء فقط على الصفوف التي يكون فيها Assigned Group من القيم المسموح بها."""
    col = _find_assigned_group_col(df)

    filtered = df[df[col].isin(ALLOWED_GROUPS)].copy()
    return filtered


def clean_organization(df: pd.DataFrame) -> pd.DataFrame:
    """تنظيف عمود Organization:
    - إزالة 'RTA\' من كل القيم
    - استبدال 'DG, Chairman of the Board Executive Directors' بـ 'Executive Affairs'
    """
    if ORG_COL not in df.columns:
        raise KeyError(f"Column '{ORG_COL}' not found in the input file.")

    result = df.copy()

    # تحويل NaN إلى فراغ قبل المعالجة النصية
    result[ORG_COL] = result[ORG_COL].fillna("")

    # إزالة RTA\ من كل القيم
    result[ORG_COL] = result[ORG_COL].astype(str).str.replace("RTA\\", "", regex=False)

    # استبدال النص المحدد بالقيمة الجديدة (استبدال تام للقيمة المطابقة)
    result[ORG_COL] = result[ORG_COL].replace(
        {"DG, Chairman of the Board Executive Directors": "Executive Affairs"}
    )

    return result


def enrich_org_from_user_file(
    df: pd.DataFrame, user_ref: pd.DataFrame
) -> tuple[pd.DataFrame, list[int]]:
    """ملء قيم Department و Organization من ملف user missing agency
    للصفوف التي يكون فيها Organization فارغ، أو تمييزها للتلوين إذا لم تُوجد.
    ترجع DataFrame محدثة وقائمة بأرقام الصفوف (0-based) التي تحتاج تلوين أصفر.
    """

    df = df.copy()

    # تأكد من الأعمدة المطلوبة في كلا الجدولين
    for col in (FIRST_NAME_COL, LAST_NAME_COL, ORG_COL):
        if col not in df.columns:
            raise KeyError(f"Column '{col}' not found in main data.")
    for col in (FIRST_NAME_COL, LAST_NAME_COL, DEPT_COL, ORG_COL):
        if col not in user_ref.columns:
            raise KeyError(f"Column '{col}' not found in user missing agency file.")

    # الصفوف ذات Organization الفارغ
    org_series = df[ORG_COL].fillna("")
    blank_mask = org_series.astype(str).str.strip().eq("")

    rows_to_color: list[int] = []
    if not blank_mask.any():
        return df, rows_to_color

    # تجهيز جدول المرجع
    ref = user_ref[[FIRST_NAME_COL, LAST_NAME_COL, DEPT_COL, ORG_COL]].copy()
    ref = ref.rename(
        columns={DEPT_COL: "Department_ref", ORG_COL: "Organization_ref"}
    )

    # العمل فقط على الصفوف التي فيها Organization فارغ
    blank_df = df[blank_mask].copy()
    blank_df = blank_df.reset_index().rename(columns={"index": "_orig_index"})

    merged = blank_df.merge(
        ref,
        on=[FIRST_NAME_COL, LAST_NAME_COL],
        how="left",
    )

    # تحديث الصفوف في df أو تمييزها للتلوين
    for row in merged.itertuples(index=False):
        orig_idx = getattr(row, "_orig_index")
        org_ref = getattr(row, "Organization_ref")
        dept_ref = getattr(row, "Department_ref")

        if pd.notna(org_ref) and str(org_ref).strip() != "":
            # تعبئة Organization من ملف المرجع
            df.at[orig_idx, ORG_COL] = org_ref
            # تعبئة Department إن وجد العمود
            if DEPT_COL in df.columns and pd.notna(dept_ref):
                df.at[orig_idx, DEPT_COL] = dept_ref
        else:
            # لا يوجد سطر مطابق في ملف user missing agency -> نلوّن لاحقاً
            rows_to_color.append(int(orig_idx))

    return df, rows_to_color


def main() -> None:
    # قراءة الملفات الخام
    open_df = pd.read_excel(OPEN_FILE)
    closed_df = pd.read_excel(CLOSED_FILE)

    # تطبيق الفلتر على عمود Assigned Group*+
    open_filtered = filter_assigned_group(open_df)
    closed_filtered = filter_assigned_group(closed_df)

    # تنظيف عمود Organization حسب المطلوب
    open_filtered = clean_organization(open_filtered).reset_index(drop=True)
    closed_filtered = clean_organization(closed_filtered).reset_index(drop=True)

    # قراءة ملف user missing agency
    user_ref_df = pd.read_excel(USER_MISSING_FILE)

    # ملء Organization/Department من ملف user missing agency أو تمييز الصفوف للتلوين
    open_filtered, open_rows_to_color = enrich_org_from_user_file(
        open_filtered, user_ref_df
    )
    closed_filtered, closed_rows_to_color = enrich_org_from_user_file(
        closed_filtered, user_ref_df
    )

    # الخطوة 10 و 11: حساب عدد التذاكر المفتوحة حسب Assigned Group من شيت Open Calls
    assigned_col_open = _find_assigned_group_col(open_filtered)
    sd_open_mask = open_filtered[assigned_col_open].isin(
        ["Service Desk", "SCCM-Nexthink Support"]
    )
    sd_open_count = int(sd_open_mask.sum())

    onsite_open_mask = open_filtered[assigned_col_open].eq("OnSite  EUS")
    onsite_open_count = int(onsite_open_mask.sum())

    # الخطوة 12: تكرار الحسابات نفسها على شيت Closed Calls
    assigned_col_closed = _find_assigned_group_col(closed_filtered)
    sd_closed_mask = closed_filtered[assigned_col_closed].isin(
        ["Service Desk", "SCCM-Nexthink Support"]
    )
    sd_closed_count = int(sd_closed_mask.sum())

    onsite_closed_mask = closed_filtered[assigned_col_closed].eq("OnSite  EUS")
    onsite_closed_count = int(onsite_closed_mask.sum())

    # إنشاء ملف جديد يحتوي على شيت لكل تقرير خام بعد الفلترة
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        open_filtered.to_excel(writer, sheet_name="Open Calls", index=False)
        closed_filtered.to_excel(writer, sheet_name="Closed Calls", index=False)

    # فتح الملف الناتج لإضافة شيت Consolidated وتلوين الخلايا الفارغة
    wb = load_workbook(OUTPUT_FILE)

    # الخطوة 6: إضافة شيت جديد باسم Consolidated (إذا لم يكن موجوداً)
    if "Consolidated" in wb.sheetnames:
        cons_ws = wb["Consolidated"]
    else:
        cons_ws = wb.create_sheet(title="Consolidated")

    # الخطوة 7: كتابة النصوص في A3 و A4
    cons_ws["A3"] = "SD Opened Ticket"
    cons_ws["A4"] = "SD Closed Ticket"

    # الخطوة 8: كتابة النصوص في A17 و A18
    cons_ws["A17"] = "Onsite Open Ticket"
    cons_ws["A18"] = "Onsite Closed Ticket"

    # الخطوة 9: كتابة النصوص في I10 و I11
    cons_ws["I10"] = "SD Closed Ticket"
    cons_ws["I11"] = "Onsite Closed Ticket"

    # الخطوة 13: ربط القيم في J10 و J11 بالقيم B4 و B18
    cons_ws["J10"] = "=B4"
    cons_ws["J11"] = "=B18"

    # الخطوة 14: جمع القيم في J10 و J11 ووضع النتيجة في J12
    cons_ws["J12"] = "=J10+J11"

    # الخطوة 15: إنشاء 2-D clustered column chart للقيم (A3,B3) و (A4,B4)
    # تنظيف أي رسوم بيانية سابقة في هذا الشيت (حتى لا تتكرر عند إعادة التشغيل)
    if hasattr(cons_ws, "_charts"):
        cons_ws._charts = []

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.style = 1  # Style 1 كما في Excel
    chart.title = "SD from 19th to 25th Feb 2026"

    # البيانات: القيم في B3:B4
    data_ref = Reference(cons_ws, min_col=2, min_row=3, max_row=4)
    # الفئات: العناوين في A3:A4 (Axis: Category / Horizontal)
    cats_ref = Reference(cons_ws, min_col=1, min_row=3, max_row=4)

    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)

    # تفعيل Data Labels مع إظهار القيمة فقط (بدون series name, category name, legend key)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showLegendKey = False

    # إخفاء خطوط الـ Major Gridlines في الـ Value Axis
    chart.y_axis.majorGridlines = chart.y_axis.majorGridlines or type("obj", (), {})()
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(
        ln=LineProperties(noFill=True)
    )

    # ضبط حجم المخطط (Plot Area) بحيث يكون العنوان واضحاً
    chart.width = 18
    chart.height = 10

    # التأكد من أن كل عمود بلون مختلف
    if chart.series:
        series = chart.series[0]
        dp0 = DataPoint(idx=0)
        dp0.graphicalProperties.solidFill = "4472C4"  # أزرق
        dp1 = DataPoint(idx=1)
        dp1.graphicalProperties.solidFill = "ED7D31"  # برتقالي
        series.dPt = [dp0, dp1]

    # إضافة الرسم البياني إلى الشيت عند موضع مناسب
    cons_ws.add_chart(chart, "D3")

    # كتابة نتيجة الخطوة 10 في الخلية B3 (عدد SD Opened Ticket من Open Calls)
    cons_ws["B3"] = sd_open_count
    # كتابة نتيجة الخطوة 11 في الخلية B17 (عدد Onsite Open Ticket من Open Calls)
    cons_ws["B17"] = onsite_open_count

    # كتابة نتائج الخطوة 12 من Closed Calls
    # العدد الأول (تكرار الخطوة 10 على Closed Calls) في الخلية B4
    cons_ws["B4"] = sd_closed_count
    # العدد الثاني (تكرار الخطوة 11 على Closed Calls) في الخلية B18
    cons_ws["B18"] = onsite_closed_count

    # تلوين الخلايا ذات Organization الفارغ والتي لم نجد لها تطابق
    if open_rows_to_color or closed_rows_to_color:
        fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # دالة مساعدة لتلوين صفوف في شيت معيّن
        def color_rows(sheet_name: str, rows_to_color: list[int]) -> None:
            if not rows_to_color:
                return
            ws = wb[sheet_name]

            # إيجاد رقم عمود Organization من الهيدر
            org_col_idx = None
            for cell in ws[HEADER_ROW]:
                if str(cell.value).strip() == ORG_COL:
                    org_col_idx = cell.column
                    break
            if org_col_idx is None:
                return

            for idx in rows_to_color:
                excel_row = DATA_START_ROW + idx  # 0-based index + تعويض الهيدر
                cell = ws.cell(row=excel_row, column=org_col_idx)
                cell.fill = fill_yellow

        color_rows("Open Calls", open_rows_to_color)
        color_rows("Closed Calls", closed_rows_to_color)

    wb.save(OUTPUT_FILE)

    print(f"تم إنشاء الملف: {OUTPUT_FILE.name}")


if __name__ == "__main__":
    main()

